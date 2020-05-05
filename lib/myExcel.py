#!/usr/bin/python3
from openpyxl import load_workbook


class MyExcel:

    def __init__(self, file_path, file_name):
        self.file_name = file_name
        self.file_path = file_path
        self.file_full_name = self.file_path + '/' + self.file_name

    # ---- 删除 str --------------------------------
    def str_del_netWithoutCom(self, strNetSrc, charSplitSrc) -> str:
        '''

        :param strNetSrc: string
        :param charSplitSrc: char
        :return:
        '''

        retStr = ''
        for eachLine in strNetSrc.split('\n'):          # 器件行
            strLineTemp = eachLine
            listLen = len(strLineTemp.split(charSplitSrc))
            if listLen > 1:
                retStr += eachLine + '\n'

        retStr = retStr.strip('\n') + '\n'

        return retStr

    # ---- 打开文件 --------------------------------
    # data_only = True  : 单元值为公式计算后的值，保存后，为值，公式丢失
    # data_only = False : 单元值为公式

    def open_file(self):
        global wb
        wb = load_workbook(self.file_full_name, data_only=True) # 打印出来是值
        # wb = load_workbook(self.file_full_name)               # 打印出来是公式

        try:
            print(wb.sheetnames)
        except Exception as e:
            print(str(e))
            return -1
        else:
            return 0

    # ---- 打印每行 --------------------------------
    #
    # def print_rows(self, sheet_name)->dict:
    def print_rows(self, sheet_name):
        global wb

        my_dict = dict()

        try:
            ws = wb.get_sheet_by_name(sheet_name)

            i = 0
            for row in ws.iter_rows():
                i += 1
                print('i = %d' % i, end='\t')

                for cell in row:
                    print(cell.value, end='\t')
                    # print('i=' + str(i) + '\t' + 'val=' + str(cell.value), end='\t')
                    # print(cell._bind_value, end='\t')
                print()
        except Exception as e:
            print(str(e))
        else:
            pass

    # ---- 获取表格到 DICT --------------------------------
    #
    # def print_rows1(self, sheet_name, start_col, stop_col) -> dict:
    def print_rows1(self, sheet_name, start_col, stop_col) -> list:
        """
        :param sheet_name: 路径及文件名
        :param start_col: 起始列
        :param stop_col: 结束列
        :return: list:[{sn:xx}, {pow:xx}, {pha:xx}, {}, ...]
        """

        global wb

        my_list = []

        try:
            ws = wb.get_sheet_by_name(sheet_name)

            my_dict = {}
            i = 0
            # for row in list(ws.iter_rows())[start_col:stop_col]:
            # print(ws.iter_rows().count)

            for row in ws.iter_rows():
                i += 1
                # print('i = %d' % i, end='\t')
                if i > 1:
                    my_dict['sn']   = row[start_col].value
                    my_dict['pow']  = row[start_col + 1].value
                    my_dict['pha']  = row[start_col + 2].value

                    my_list.append(my_dict.copy())
                    my_dict.clear()

                    # print('i=' + str(i) + '\t' + 'val=' + str(cell.value), end='\t')
                    # print(cell._bind_value, end='\t')

            # print(my_list)

        except Exception as e:
            print(str(e))
            return -1
        else:
            return my_list

